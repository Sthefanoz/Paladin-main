from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import F
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST

from activos.models import Activo
from .models import Amenaza, Vulnerabilidad, Control, Riesgo
from .forms import (
    AmenazaForm, VulnerabilidadForm, ControlForm, RiesgoForm,
)


# Rangos de criticidad (coherentes con Riesgo._clasificar): 1-25.
RANGOS_CRITICIDAD = {
    'Bajo': (1, 4),
    'Medio': (5, 9),
    'Alto': (10, 15),
    'Crítico': (16, 25),
}


# ---------------------------------------------------------------------------
# Riesgos (CRUD) — módulos: identificación, valoración, tratamiento, residual
# ---------------------------------------------------------------------------

def lista_riesgos(request):
    """Lista de riesgos con filtro por criticidad/estado y orden configurable."""
    riesgos = (
        Riesgo.objects
        .select_related('activo', 'amenaza', 'vulnerabilidad')
        # nivel inherente calculado en la BD para poder ordenar/filtrar
        .annotate(nivel=F('probabilidad') * F('impacto'))
    )

    nivel_sel = request.GET.get('nivel', '')
    estado_sel = request.GET.get('estado', '')
    orden = request.GET.get('orden', 'criticidad')  # criticidad | reciente

    if nivel_sel in RANGOS_CRITICIDAD:
        lo, hi = RANGOS_CRITICIDAD[nivel_sel]
        riesgos = riesgos.filter(nivel__gte=lo, nivel__lte=hi)

    if estado_sel:
        riesgos = riesgos.filter(estado=estado_sel)

    if orden == 'reciente':
        riesgos = riesgos.order_by('-fecha_actualizacion')
    else:
        riesgos = riesgos.order_by('-nivel', '-fecha_actualizacion')

    return render(request, 'riesgos/lista_riesgos.html', {
        'riesgos': riesgos,
        'niveles': RANGOS_CRITICIDAD.keys(),
        'estados': Riesgo.ESTADOS,
        'nivel_sel': nivel_sel,
        'estado_sel': estado_sel,
        'orden': orden,
    })


def crear_riesgo(request):
    form = RiesgoForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_riesgos')
    return render(
        request, 'riesgos/form_riesgo.html',
        {'form': form, 'titulo': 'Registrar Riesgo'}
    )


def editar_riesgo(request, id):
    riesgo = get_object_or_404(Riesgo, id=id)
    form = RiesgoForm(request.POST or None, instance=riesgo)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_riesgos')
    return render(
        request, 'riesgos/form_riesgo.html',
        {'form': form, 'titulo': 'Editar Riesgo'}
    )


def detalle_riesgo(request, id):
    riesgo = get_object_or_404(
        Riesgo.objects.select_related('activo', 'amenaza', 'vulnerabilidad'),
        id=id,
    )
    return render(request, 'riesgos/detalle_riesgo.html', {'riesgo': riesgo})


def eliminar_riesgo(request, id):
    riesgo = get_object_or_404(Riesgo, id=id)
    if request.method == 'POST':
        riesgo.delete()
        return redirect('lista_riesgos')
    return render(request, 'riesgos/confirmar_eliminar.html',
                  {'objeto': riesgo, 'volver': 'lista_riesgos'})


# ---------------------------------------------------------------------------
# Creación "al vuelo" (AJAX) desde el formulario de riesgo
# ---------------------------------------------------------------------------

@require_POST
def crear_amenaza_ajax(request):
    """Crea una amenaza desde el modal y devuelve su id/texto en JSON."""
    form = AmenazaForm(request.POST)
    if form.is_valid():
        obj = form.save()
        return JsonResponse({'ok': True, 'id': obj.id, 'text': str(obj)})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


@require_POST
def crear_vulnerabilidad_ajax(request):
    """Crea una vulnerabilidad desde el modal y devuelve su id/texto en JSON."""
    form = VulnerabilidadForm(request.POST)
    if form.is_valid():
        obj = form.save()
        return JsonResponse({'ok': True, 'id': obj.id, 'text': str(obj)})
    return JsonResponse({'ok': False, 'errors': form.errors}, status=400)


# ---------------------------------------------------------------------------
# Catálogos: Amenazas, Vulnerabilidades, Controles
# ---------------------------------------------------------------------------

def _eliminar_catalogo(request, modelo, id, volver, en_uso):
    """Elimina un elemento de catálogo, bloqueando si está en uso por un riesgo."""
    obj = get_object_or_404(modelo, id=id)
    bloqueado = en_uso(obj)
    if request.method == 'POST' and not bloqueado:
        obj.delete()
        return redirect(volver)
    mensaje = None
    if bloqueado:
        mensaje = ('No se puede eliminar porque está siendo utilizado por uno o '
                   'más riesgos. Quite primero esa relación en los riesgos.')
    return render(request, 'riesgos/confirmar_eliminar.html',
                  {'objeto': obj, 'volver': volver, 'bloqueado': bloqueado,
                   'mensaje': mensaje})


def lista_amenazas(request):
    return render(request, 'riesgos/lista_amenazas.html',
                  {'amenazas': Amenaza.objects.all()})


def crear_amenaza(request):
    form = AmenazaForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_amenazas')
    return render(request, 'riesgos/form_simple.html',
                  {'form': form, 'titulo': 'Registrar Amenaza',
                   'volver': 'lista_amenazas'})


def eliminar_amenaza(request, id):
    return _eliminar_catalogo(
        request, Amenaza, id, 'lista_amenazas',
        en_uso=lambda o: o.riesgo_set.exists(),
    )


def lista_vulnerabilidades(request):
    return render(request, 'riesgos/lista_vulnerabilidades.html',
                  {'vulnerabilidades': Vulnerabilidad.objects.all()})


def eliminar_vulnerabilidad(request, id):
    return _eliminar_catalogo(
        request, Vulnerabilidad, id, 'lista_vulnerabilidades',
        en_uso=lambda o: o.riesgo_set.exists(),
    )


def crear_vulnerabilidad(request):
    form = VulnerabilidadForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_vulnerabilidades')
    return render(request, 'riesgos/form_simple.html',
                  {'form': form, 'titulo': 'Registrar Vulnerabilidad',
                   'volver': 'lista_vulnerabilidades'})


def lista_controles(request):
    return render(request, 'riesgos/lista_controles.html',
                  {'controles': Control.objects.all()})


def crear_control(request):
    form = ControlForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('lista_controles')
    return render(request, 'riesgos/form_simple.html',
                  {'form': form, 'titulo': 'Registrar Control',
                   'volver': 'lista_controles'})


def eliminar_control(request, id):
    return _eliminar_catalogo(
        request, Control, id, 'lista_controles',
        en_uso=lambda o: o.riesgos.exists(),
    )


# ---------------------------------------------------------------------------
# Monitoreo y supervisión (panel) + Comunicación (reporte)
# ---------------------------------------------------------------------------

def _resumen_riesgos():
    """Construye las métricas que comparten el dashboard y el reporte."""
    riesgos = list(
        Riesgo.objects
        .select_related('activo', 'amenaza', 'vulnerabilidad')
        .prefetch_related('controles_propuestos')
        .all()
    )
    # Ordenados por criticidad para que el panel/reporte muestren lo importante.
    riesgos.sort(key=lambda r: r.nivel_riesgo, reverse=True)

    por_estado = {clave: 0 for clave, _ in Riesgo.ESTADOS}
    por_nivel = {'Bajo': 0, 'Medio': 0, 'Alto': 0, 'Crítico': 0}
    por_estrategia = {clave: 0 for clave, _ in Riesgo.ESTRATEGIAS}
    por_estrategia['(sin tratamiento)'] = 0

    # Métricas de evolución (inherente -> residual).
    suma_inherente = 0
    suma_residual = 0
    con_residual = 0

    for r in riesgos:
        por_estado[r.estado] = por_estado.get(r.estado, 0) + 1
        por_nivel[r.nivel_riesgo_texto] = por_nivel.get(r.nivel_riesgo_texto, 0) + 1
        clave_est = r.estrategia if r.estrategia else '(sin tratamiento)'
        por_estrategia[clave_est] = por_estrategia.get(clave_est, 0) + 1
        if r.riesgo_residual is not None:
            suma_inherente += r.nivel_riesgo
            suma_residual += r.riesgo_residual
            con_residual += 1

    etiquetas_estado = dict(Riesgo.ESTADOS)
    estados = [
        {'nombre': etiquetas_estado[k], 'total': v}
        for k, v in por_estado.items()
    ]

    etiquetas_estr = dict(Riesgo.ESTRATEGIAS)
    estrategias = [
        {'nombre': etiquetas_estr.get(k, k), 'total': v}
        for k, v in por_estrategia.items() if v
    ]

    # Evolución global del riesgo gracias a los tratamientos.
    reduccion_global = None
    if con_residual and suma_inherente:
        reduccion_global = round((suma_inherente - suma_residual) / suma_inherente * 100)

    return {
        'total': len(riesgos),
        'estados': estados,
        'estrategias': estrategias,
        'niveles': por_nivel,
        'criticos': riesgos[:10],
        'riesgos': riesgos,
        'suma_inherente': suma_inherente,
        'suma_residual': suma_residual,
        'reduccion_global': reduccion_global,
        'con_residual': con_residual,
    }


def dashboard(request):
    contexto = _resumen_riesgos()
    contexto['total_activos'] = Activo.objects.count()
    return render(request, 'riesgos/dashboard.html', contexto)


def reporte(request):
    """Reporte para partes interesadas (comunicación y consulta)."""
    contexto = _resumen_riesgos()
    return render(request, 'riesgos/reporte.html', contexto)


def reporte_excel(request):
    """Genera el informe de riesgos en un archivo Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    riesgos = (
        Riesgo.objects
        .select_related('activo', 'amenaza', 'vulnerabilidad')
        .prefetch_related('controles_propuestos')
        .all()
    )
    riesgos = sorted(riesgos, key=lambda r: r.nivel_riesgo, reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Riesgos'

    encabezados = [
        'Activo', 'Tipo', 'Amenaza', 'Vulnerabilidad', 'CVE',
        'Probabilidad', 'Impacto', 'Riesgo inherente', 'Nivel',
        'Estrategia', 'Controles (ISO 27002)', 'Responsable',
        'Prob. residual', 'Impacto residual', 'Riesgo residual', 'Nivel residual',
        'Reducción %', 'Estado', 'Fecha tratamiento', 'Fecha control',
        'Observaciones',
    ]
    ws.append(encabezados)

    # Estilo del encabezado.
    cabecera_fill = PatternFill('solid', fgColor='212529')
    cabecera_font = Font(color='FFFFFF', bold=True)
    for celda in ws[1]:
        celda.fill = cabecera_fill
        celda.font = cabecera_font
        celda.alignment = Alignment(horizontal='center', vertical='center')

    colores_nivel = {
        'Bajo': '198754', 'Medio': 'FFC107', 'Alto': 'FD7E14', 'Crítico': 'DC3545',
    }

    for r in riesgos:
        controles = ', '.join(c.codigo for c in r.controles_propuestos.all())
        ws.append([
            r.activo.nombre, r.activo.get_tipo_display(),
            r.amenaza.nombre, r.vulnerabilidad.nombre, r.vulnerabilidad.cve or '',
            r.probabilidad, r.impacto, r.nivel_riesgo, r.nivel_riesgo_texto,
            r.get_estrategia_display() if r.estrategia else '',
            controles, r.responsable or '',
            r.probabilidad_residual or '', r.impacto_residual or '',
            r.riesgo_residual if r.riesgo_residual is not None else '',
            r.riesgo_residual_texto or '',
            r.porcentaje_reduccion if r.porcentaje_reduccion is not None else '',
            r.get_estado_display(),
            r.fecha_tratamiento.isoformat() if r.fecha_tratamiento else '',
            r.fecha_control.isoformat() if r.fecha_control else '',
            r.observaciones or '',
        ])
        # Colorear la celda del nivel inherente según criticidad.
        celda_nivel = ws.cell(row=ws.max_row, column=9)
        color = colores_nivel.get(r.nivel_riesgo_texto)
        if color:
            celda_nivel.fill = PatternFill('solid', fgColor=color)
            celda_nivel.font = Font(bold=True, color='FFFFFF')

    # Ancho de columnas aproximado.
    anchos = [22, 12, 22, 24, 16, 12, 9, 15, 10, 14, 22, 16,
              12, 14, 14, 14, 11, 14, 16, 14, 30]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho
    ws.freeze_panes = 'A2'

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="reporte_riesgos.xlsx"'
    wb.save(resp)
    return resp
